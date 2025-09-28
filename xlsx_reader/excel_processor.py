"""Excel processing module for reading XLSX files.

Provides functions to:
- list sheet names
- count data rows per sheet (excluding header row)
- process an entire workbook with a progress callback
"""

from __future__ import annotations

from collections.abc import Callable
from pathlib import Path

from openpyxl import load_workbook


def _ensure_exists(file_path: str) -> None:
    """Raise FileNotFoundError if the path does not exist."""
    if not Path(file_path).exists():
        raise FileNotFoundError(file_path)


def get_sheet_names(file_path: str) -> list[str]:
    """Return sheet names in creation order.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        list[str]: Sheet names.
    """
    _ensure_exists(file_path)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def get_sheet_row_count(file_path: str, sheet_name: str) -> int:
    """Return number of non-empty data rows in a sheet (excluding header row).

    Args:
        file_path: Path to the .xlsx file.
        sheet_name: Name of the sheet to inspect.

    Returns:
        int: Count of rows that contain any data, starting from row 2.
    """
    _ensure_exists(file_path)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet {sheet_name!r} not found in workbook.")
        ws = wb[sheet_name]

        count = 0
        # Skip header row (row=1); count any row with at least one non-empty cell.
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None and cell != "" for cell in row):
                count += 1
        return count
    finally:
        wb.close()


def process_excel_file(
    file_path: str,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> dict[str, int]:
    """Process all sheets and return a mapping {sheet_name: row_count}.

    The progress callback (if provided) is called as:
        progress_callback(current_index, total_sheets, sheet_name)
    where current_index is zero-based.

    Args:
        file_path: Path to the .xlsx file.
        progress_callback: Optional progress reporter.

    Returns:
        dict[str, int]: Row counts per sheet.
    """
    _ensure_exists(file_path)

    sheets = get_sheet_names(file_path)
    total = len(sheets)
    results: dict[str, int] = {}

    for idx, name in enumerate(sheets):  # idx = 0..total-1
        if progress_callback:
            progress_callback(idx, total, name)
        results[name] = get_sheet_row_count(file_path, name)

    return results
