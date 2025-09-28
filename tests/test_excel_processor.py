"""Tests for Excel processing functions.

Covers:
- Reading sheet names
- Counting data rows (excluding header)
- Processing all sheets with optional progress callback
- File-not-found handling
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from xlsx_reader.excel_processor import (
    get_sheet_names,
    get_sheet_row_count,
    process_excel_file,
)


def _create_test_excel(path: Path) -> None:
    """Create a simple test workbook with 3 sheets."""
    wb = Workbook()
    # remove default "Sheet"
    wb.remove(wb.active)

    # Sheet1: 10 data rows (+1 header row)
    s1 = wb.create_sheet("Sheet1")
    s1["A1"], s1["B1"] = "A", "B"
    for i in range(1, 11):
        s1[f"A{i + 1}"] = i
        s1[f"B{i + 1}"] = f"Row {i}"

    # Sheet2: 5 data rows (+1 header row)
    s2 = wb.create_sheet("Sheet2")
    s2["X1"], s2["Y1"] = "X", "Y"
    for i in range(1, 6):
        s2[f"X{i + 1}"] = i
        s2[f"Y{i + 1}"] = f"Data {i}"

    # EmptySheet: headers only (0 data rows)
    s3 = wb.create_sheet("EmptySheet")
    s3["A1"], s3["B1"] = "Col1", "Col2"

    wb.save(path)


@pytest.fixture
def test_excel_file(tmp_path: Path) -> str:
    """Provide a temporary .xlsx file populated with predictable data."""
    xlsx_path = tmp_path / "sample.xlsx"
    _create_test_excel(xlsx_path)
    return str(xlsx_path)


class TestExcelProcessor:
    def test_get_sheet_names(self, test_excel_file: str) -> None:
        names = get_sheet_names(test_excel_file)
        assert isinstance(names, list)
        assert set(names) == {"Sheet1", "Sheet2", "EmptySheet"}
        # preserve creation order
        assert names == ["Sheet1", "Sheet2", "EmptySheet"]

    @pytest.mark.parametrize(
        ("sheet", "expected_rows"),
        [
            ("Sheet1", 10),
            ("Sheet2", 5),
            ("EmptySheet", 0),
        ],
    )
    def test_get_sheet_row_count(
        self, test_excel_file: str, sheet: str, expected_rows: int
    ) -> None:
        assert get_sheet_row_count(test_excel_file, sheet) == expected_rows

    def test_process_excel_file(self, test_excel_file: str) -> None:
        results: dict[str, int] = process_excel_file(test_excel_file)
        assert results == {"Sheet1": 10, "Sheet2": 5, "EmptySheet": 0}

    def test_process_excel_file_with_callback(self, test_excel_file: str) -> None:
        calls: list[tuple[int, int, str]] = []

        def cb(current: int, total: int, sheet_name: str) -> None:
            calls.append((current, total, sheet_name))

        process_excel_file(test_excel_file, progress_callback=cb)

        # one callback per sheet, in order
        assert calls == [
            (0, 3, "Sheet1"),
            (1, 3, "Sheet2"),
            (2, 3, "EmptySheet"),
        ]

    def test_file_not_found(self) -> None:
        with pytest.raises(FileNotFoundError):
            get_sheet_names("nonexistent.xlsx")
        with pytest.raises(FileNotFoundError):
            get_sheet_row_count("nonexistent.xlsx", "Sheet1")
        with pytest.raises(FileNotFoundError):
            process_excel_file("nonexistent.xlsx")
